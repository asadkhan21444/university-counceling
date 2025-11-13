from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.http import JsonResponse,HttpResponse
from django.contrib.auth import login
import json
from django.views.decorators.csrf import csrf_exempt
from django.contrib import messages
from django.views.decorators.csrf import csrf_protect
from django.contrib.auth.decorators import login_required
from django.views.decorators.http import require_POST
from django.db import transaction
from decimal import Decimal, InvalidOperation
from django.db import IntegrityError
from django.views.decorators.http import require_http_methods
from django.db.models import Q
import openpyxl
from django.utils import timezone
from openpyxl.utils import get_column_letter
from django.db.models import Sum
from django.core.paginator import Paginator
from datetime import datetime, timedelta
from django.template.loader import get_template
from django.views.decorators.http import require_POST, require_GET
from django.views.generic import TemplateView
from django.contrib.auth.decorators import login_required, permission_required
import openpyxl
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from django.contrib.contenttypes.models import ContentType
from django.contrib.auth.models import Permission, Group
from django.core.exceptions import PermissionDenied
from django.db.models import Avg, Count
#####################################################
from .models import(
    CustomUser,
    Student,
    Counsellor,
    Appointment,
    CounsellorAvailability,
    CounsellingSession,
    SessionDocument,
    EmergencyContact,
    CounsellorReview
)
########################################
from .forms import (
    CustomUserCreationForm,
    CustomAuthenticationForm,
    StudentForm,
    CounsellorForm,
    AppointmentForm,
    CounsellorAvailabilityForm,
    SessionDocumentForm,
    EmergencyContactForm
)
from django import forms
from django.utils import timezone
from datetime import datetime, timedelta
#####################################################################################################################################################
#####################################################################################################################################################
def home(request):
    return render(request, 'My_Home/index.html')
#####################################################################################################################################################
#####################################################################################################################################################
class ServicesView(TemplateView):
    template_name = 'My_Home/services.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Add any additional context data for services page
        context['services'] = [
            {
                'id': 'student-counselling',
                'title': 'Student Counselling',
                'icon': 'user-graduate',
                'color': 'primary',
                'description': 'Comprehensive support including academic guidance, career counseling, and personal development to help you thrive in your university journey.',
                'features': [
                    'Academic performance monitoring and improvement strategies',
                    'Career path guidance and exploration',
                    'Personal development and growth sessions',
                    'Stress management and coping techniques',
                    'Time management and study skills development',
                    'Goal setting and achievement planning'
                ],
                'badge': 'Available to all students',
                'button_text': 'Schedule Session'
            },
            {
                'id': 'counsellor-guidance',
                'title': 'Counsellor Guidance',
                'icon': 'chalkboard-teacher',
                'color': 'info',
                'description': 'Connect with experienced counsellors, view their profiles, and choose the right professional for your specific needs and preferences.',
                'features': [
                    'Detailed counsellor profiles and specialties',
                    'Expertise matching based on your needs',
                    'Professional qualifications and experience verification',
                    'Real-time availability and scheduling',
                    'Student reviews and ratings',
                    'Specialized support for diverse needs'
                ],
                'badge': 'Personalized matching',
                'button_text': 'Find a Counsellor'
            },
            {
                'id': 'appointments',
                'title': 'Appointments',
                'icon': 'calendar-alt',
                'color': 'warning',
                'description': 'Easy-to-use scheduling system for booking, managing, and tracking your counselling sessions with real-time availability.',
                'features': [
                    'Intuitive online booking system',
                    'Automated session reminders via email and SMS',
                    'Flexible rescheduling and cancellation options',
                    'Calendar integration with popular platforms',
                    'Waitlist for fully booked counsellors',
                    'Recurring appointment options'
                ],
                'badge': 'Easy scheduling',
                'button_text': 'Book Appointment'
            },
            {
                'id': 'reports',
                'title': 'Reports & Documents',
                'icon': 'file-alt',
                'color': 'success',
                'description': 'Secure document management system for session notes, progress reports, forms, and other important counselling documentation.',
                'features': [
                    'Secure storage of session notes and summaries',
                    'Progress tracking and milestone documentation',
                    'Safe document sharing between students and counsellors',
                    'Role-based access control for privacy',
                    'Export options for personal records',
                    'Automated report generation'
                ],
                'badge': 'Secure storage',
                'button_text': 'Access Documents'
            }
        ]
        
        context['process_steps'] = [
            {
                'number': 1,
                'title': 'Register',
                'description': 'Create your student account and complete your profile'
            },
            {
                'number': 2,
                'title': 'Browse',
                'description': 'Explore counsellor profiles and find the right match'
            },
            {
                'number': 3,
                'title': 'Schedule',
                'description': 'Book appointments at your convenience'
            },
            {
                'number': 4,
                'title': 'Connect',
                'description': 'Attend sessions and track your progress'
            }
        ]
        return context
#####################################################################################################################################################
#####################################################################################################################################################
class AboutView(TemplateView):
    template_name = 'My_Home/aboutus.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Add context data for about page
        context['mission'] = {
            'title': 'Our Mission',
            'description': 'To empower students to achieve academic excellence and personal growth through accessible, professional counselling services that address their unique challenges and aspirations.',
            'icon': 'rocket'
        }
        
        context['vision'] = {
            'title': 'Our Vision',
            'description': 'To create a university environment where every student has access to comprehensive support systems that foster holistic development, resilience, and lifelong success.',
            'icon': 'eye'
        }
        
        context['values'] = [
            {
                'title': 'Confidentiality',
                'description': 'We maintain the highest standards of privacy and confidentiality in all our interactions with students.',
                'icon': 'user-shield'
            },
            {
                'title': 'Empathy',
                'description': 'We approach every student with understanding, compassion, and a genuine desire to help.',
                'icon': 'hands-helping'
            },
            {
                'title': 'Excellence',
                'description': 'We are committed to providing the highest quality counselling services through continuous improvement.',
                'icon': 'graduation-cap'
            }
        ]
        
        context['team'] = [
            {
                'name': 'Dr. Sarah Johnson',
                'position': 'Director of Counselling Services',
                'description': 'With over 15 years of experience in student counselling, Dr. Johnson leads our team with expertise and compassion.',
                'image': 'https://images.unsplash.com/photo-1560250097-0b93528c311a?ixlib=rb-4.0.3&auto=format&fit=crop&w=400&q=80'
            },
            {
                'name': 'Michael Chen',
                'position': 'Head of Academic Counselling',
                'description': 'Michael specializes in academic performance strategies and has helped thousands of students improve their grades.',
                'image': 'https://images.unsplash.com/photo-1551836026-d5c88ac5d691?ixlib=rb-4.0.3&auto=format&fit=crop&w=400&q=80'
            },
            {
                'name': 'Dr. Maria Rodriguez',
                'position': 'Lead Career Counsellor',
                'description': 'Dr. Rodriguez has extensive experience in career development and has guided countless students toward fulfilling careers.',
                'image': 'https://images.unsplash.com/photo-1580489944761-15a19d654956?ixlib=rb-4.0.3&auto=format&fit=crop&w=400&q=80'
            }
        ]
        
        context['stats'] = [
            {
                'value': '5,000+',
                'label': 'Students Supported'
            },
            {
                'value': '98%',
                'label': 'Satisfaction Rate'
            },
            {
                'value': '50+',
                'label': 'Qualified Counsellors'
            },
            {
                'value': '24/7',
                'label': 'Support Availability'
            }
        ]
        
        context['testimonials'] = [
            {
                'name': 'Emily Johnson',
                'program': 'Psychology Major, Class of 2023',
                'text': '"The counselling services completely transformed my university experience. I went from struggling academically to achieving honors with their guidance."',
                'image': 'https://images.unsplash.com/photo-1494790108755-2616b612b786?ixlib=rb-4.0.3&auto=format&fit=crop&w=100&q=80',
                'rating': 5
            },
            {
                'name': 'James Wilson',
                'program': 'Business Administration, Class of 2024',
                'text': '"I was unsure about my career path until I met with a career counsellor. They helped me discover my strengths and find the perfect internship opportunity."',
                'image': 'https://images.unsplash.com/photo-1507003211169-0a1dd7228f2d?ixlib=rb-4.0.3&auto=format&fit=crop&w=100&q=80',
                'rating': 5
            },
            {
                'name': 'Sophia Martinez',
                'program': 'Engineering, Class of 2023',
                'text': '"The stress management techniques I learned through counselling helped me balance my coursework and extracurricular activities without burning out."',
                'image': 'https://images.unsplash.com/photo-1438761681033-6461ffad8d80?ixlib=rb-4.0.3&auto=format&fit=crop&w=100&q=80',
                'rating': 5
            }
        ]
        
        return context
#####################################################################################################################################################
#####################################################################################################################################################
# Function-based view alternatives (if you prefer these instead)
def home_view(request):
    return render(request, 'My_Home/index.html')
#####################################################################################################################################################
#####################################################################################################################################################
def services_view(request):
    context = {
        'services': [
            # ... same services data as above
        ],
        'process_steps': [
            # ... same process steps data as above
        ]
    }
    return render(request, 'My_Home/services.html', context)
#####################################################################################################################################################
#####################################################################################################################################################
def about_view(request):
    context = {
        'mission': {
            # ... same mission data as above
        },
        'vision': {
            # ... same vision data as above
        },
        'values': [
            # ... same values data as above
        ],
        'team': [
            # ... same team data as above
        ],
        'stats': [
            # ... same stats data as above
        ],
        'testimonials': [
            # ... same testimonials data as above
        ]
    }
    return render(request, 'My_Home/aboutus.html', context)
#####################################################################################################################################################
#####################################################################################################################################################
def system_user_home(request):
    """Main landing page for user authentication"""
    return render(request, 'My_Home/System_user.html')
#####################################################################################################################################################
#####################################################################################################################################################
def register_view(request):
    # Check if admin already exists and hide admin role from choices
    admin_exists = CustomUser.objects.filter(role='admin').exists()
    
    if request.method == 'POST':
        form = CustomUserCreationForm(request.POST, request.FILES, admin_exists=admin_exists)
        if form.is_valid():
            try:
                user = form.save()
                
                # Set additional permissions for admin users
                if user.role == 'admin':
                    user.is_staff = True
                    user.is_superuser = True
                    user.save()
                
                messages.success(request, 'Registration successful! Please login.')
                return redirect('login')
            except Exception as e:
                messages.error(request, f'Error during registration: {str(e)}')
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = CustomUserCreationForm(admin_exists=admin_exists)
    
    return render(request, 'User/register.html', {
        'form': form,
        'admin_exists': admin_exists
    })
#####################################################################################################################################################
#####################################################################################################################################################
def login_view(request):
    if request.method == 'POST':
        form = CustomAuthenticationForm(request, data=request.POST)
        if form.is_valid():
            user = form.get_user()  # cleaner way
            login(request, user)
            messages.success(request, f'Welcome back, {user.username}!')

            # Redirect based on role
            if user.role == 'admin':
                return redirect('admin_dashboard')
            elif user.role == 'counsellor':
                return redirect('counsellor_dashboard')
            elif user.role == 'student':
                return redirect('student_dashboard')
            else:
                return redirect('home')  # fallback
        else:
            messages.error(request, 'Invalid username or password.')
    else:
        form = CustomAuthenticationForm()

    return render(request, 'User/login.html', {'form': form})
#####################################################################################################################################################
#####################################################################################################################################################
def logout_view(request):
    logout(request)
    messages.success(request, 'You have been successfully logged out.')
    return redirect('login')
#####################################################################################################################################################
#####################################################################################################################################################
@login_required
def admin_dashboard(request):
    if not request.user.is_admin():
        messages.error(request, 'Access denied. Admin privileges required.')
        return redirect('login')
    
    try:
        # Get dashboard statistics with fallbacks
        total_users = CustomUser.objects.count() or 0
        total_students = CustomUser.objects.filter(role='student').count() or 0
        total_counsellors = CustomUser.objects.filter(role='counsellor').count() or 0
        total_admins = CustomUser.objects.filter(role='admin').count() or 0
        
        # Get user data for tables
        all_users = CustomUser.objects.all().order_by('-date_joined')[:50]
        students = CustomUser.objects.filter(role='student').order_by('-date_joined')[:10]
        counsellors = CustomUser.objects.filter(role='counsellor').order_by('-date_joined')[:10]
        admins = CustomUser.objects.filter(role='admin').order_by('-date_joined')[:10]
        
        # Get today's date for new users calculation
        today = timezone.now().date()
        new_users_today = CustomUser.objects.filter(date_joined__date=today).count() or 0
        
        # Get active users count
        active_users_count = CustomUser.objects.filter(is_active=True).count() or 0
        inactive_users_count = max(0, total_users - active_users_count)
        
        # Generate user growth data for last 7 days
        growth_data = []
        growth_labels = []
        
        for i in range(6, -1, -1):
            date = today - timedelta(days=i)
            count = CustomUser.objects.filter(date_joined__date=date).count() or 0
            growth_data.append(count)
            growth_labels.append(date.strftime("%b %d"))
        
        # Additional statistics
        active_users_today = CustomUser.objects.filter(last_login__date=today).count() or 0
        
        context = {
            'total_users': total_users,
            'total_students': total_students,
            'total_counsellors': total_counsellors,
            'total_admins': total_admins,
            'all_users': all_users,
            'students': students,
            'counsellors': counsellors,
            'admins': admins,
            'new_users_today': new_users_today,
            'active_users': active_users_today,
            'active_users_count': active_users_count,
            'inactive_users_count': inactive_users_count,
            'pending_approvals': 0,
            'total_sessions': 0,
            'growth_data': json.dumps(growth_data),
            'growth_labels': json.dumps(growth_labels),
        }
        
    except Exception as e:
        # Fallback context if there's any error
        print(f"Error in admin_dashboard: {e}")
        context = {
            'total_users': 0,
            'total_students': 0,
            'total_counsellors': 0,
            'total_admins': 0,
            'all_users': [],
            'students': [],
            'counsellors': [],
            'admins': [],
            'new_users_today': 0,
            'active_users': 0,
            'active_users_count': 0,
            'inactive_users_count': 0,
            'pending_approvals': 0,
            'total_sessions': 0,
            'growth_data': json.dumps([0, 0, 0, 0, 0, 0, 0]),
            'growth_labels': json.dumps(['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun']),
        }
    
    return render(request, 'My_Admin/admin_dashboard.html', context)
#####################################################(For Student)################################################################################################
#####################################################(For Student))###################################################################################
@login_required
def student_list(request):
    students = Student.objects.all()
    return render(request, 'My_Admin/Student_Management/student_list.html', {'students': students})

@login_required
def student_add(request):
    if request.method == "POST":
        form = StudentForm(request.POST)
        if form.is_valid():
            student = form.save(commit=False)
            student.user = request.user  # Assign current user if needed
            student.save()
            messages.success(request, "Student added successfully!")
            return redirect('student_list')
        else:
            messages.error(request, "Error adding student. Please check the form.")
    else:
        form = StudentForm()
    return render(request, 'My_Admin/Student_Management/student_form.html', {'form': form, 'title': 'Add Student'})

@login_required
def student_update(request, pk):
    student = get_object_or_404(Student, pk=pk)
    if request.method == "POST":
        form = StudentForm(request.POST, instance=student)
        if form.is_valid():
            form.save()
            messages.success(request, "Student updated successfully!")
            return redirect('student_list')
        else:
            messages.error(request, "Error updating student.")
    else:
        form = StudentForm(instance=student)
    return render(request, 'My_Admin/Student_Management/student_form.html', {'form': form, 'title': 'Update Student'})

@login_required
def student_delete(request, pk):
    student = get_object_or_404(Student, pk=pk)
    if request.method == "POST":
        student.delete()
        messages.success(request, "Student deleted successfully!")
        return redirect('student_list')
    return JsonResponse({'error': 'Invalid request'}, status=400)
#####################################################(For Counseller)################################################################################################
#####################################################(For Counseller))###################################################################################
# ✅ View all Counsellors
def counsellor_list(request):
    counsellors = Counsellor.objects.all()
    return render(request, "My_Admin/Counsellor_Management/counsellor_list.html", {"counsellors": counsellors})

# ✅ Add a Counsellor
def counsellor_add(request):
    if request.method == "POST":
        form = CounsellorForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Counsellor added successfully!")
            return redirect("counsellor_list")
        else:
            messages.error(request, "Failed to add counsellor. Please check the form.")
    else:
        form = CounsellorForm()
    return render(request, "My_Admin/Counsellor_Management/counsellor_form.html", {"form": form, "title": "Add Counsellor"})

# ✅ Update Counsellor
def counsellor_update(request, pk):
    counsellor = get_object_or_404(Counsellor, pk=pk)
    if request.method == "POST":
        form = CounsellorForm(request.POST, instance=counsellor)
        if form.is_valid():
            form.save()
            messages.success(request, "Counsellor updated successfully!")
            return redirect("counsellor_list")
        else:
            messages.error(request, "Failed to update counsellor.")
    else:
        form = CounsellorForm(instance=counsellor)
    return render(request, "My_Admin/Counsellor_Management/counsellor_form.html", {"form": form, "title": "Update Counsellor"})

# ✅ Delete Counsellor (with SweetAlert confirmation via AJAX)
@csrf_exempt
def counsellor_delete(request, pk):
    if request.method == "POST":
        try:
            counsellor = get_object_or_404(Counsellor, pk=pk)
            counsellor.delete()
            return JsonResponse({"status": "success", "message": "Counsellor deleted successfully!"})
        except Exception as e:
            return JsonResponse({"status": "error", "message": str(e)})
    return JsonResponse({"status": "error", "message": "Invalid request"})
#####################################################(For Appointment)################################################################################################
#####################################################(For Appointment))###################################################################################
# View all appointments
def appointment_list(request):
    appointments = Appointment.objects.all().select_related('student', 'counsellor')
    return render(request, 'My_Admin/appointments/list.html', {'appointments': appointments})

# Add new appointment
def appointment_create(request):
    if request.method == "POST":
        form = AppointmentForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Appointment created successfully!")
            return redirect('appointment_list')
        else:
            messages.error(request, "Error creating appointment. Please check the form.")
    else:
        form = AppointmentForm()
    return render(request, 'My_Admin/appointments/form.html', {'form': form, 'title': 'Add Appointment'})

# Update appointment
def appointment_update(request, pk):
    appointment = get_object_or_404(Appointment, pk=pk)
    if request.method == "POST":
        form = AppointmentForm(request.POST, instance=appointment)
        if form.is_valid():
            form.save()
            messages.success(request, "Appointment updated successfully!")
            return redirect('appointment_list')
        else:
            messages.error(request, "Error updating appointment.")
    else:
        form = AppointmentForm(instance=appointment)
    return render(request, 'My_Admin/appointments/form.html', {'form': form, 'title': 'Update Appointment'})

# Delete appointment (with SweetAlert confirmation)
@csrf_exempt
def appointment_delete(request, pk):
    appointment = get_object_or_404(Appointment, pk=pk)
    appointment.delete()
    return JsonResponse({"success": True})
#####################################################(For counseller_availaablity)################################################################################################
#####################################################(For counseller_availaablity)###################################################################################
# View all records
def availability_list(request):
    availabilities = CounsellorAvailability.objects.select_related('counsellor').all()
    return render(request, "My_Admin/CounsellorAvailability/availability_list.html", {"availabilities": availabilities})


# Add record
def availability_add(request):
    if request.method == "POST":
        form = CounsellorAvailabilityForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Availability added successfully!")
            return redirect("availability_list")
        else:
            messages.error(request, "Error adding availability. Please check the form.")
    else:
        form = CounsellorAvailabilityForm()
    return render(request, "My_Admin/CounsellorAvailability/availability_form.html", {"form": form, "title": "Add Availability"})


# Update record
def availability_update(request, pk):
    availability = get_object_or_404(CounsellorAvailability, pk=pk)
    if request.method == "POST":
        form = CounsellorAvailabilityForm(request.POST, instance=availability)
        if form.is_valid():
            form.save()
            messages.success(request, "Availability updated successfully!")
            return redirect("availability_list")
        else:
            messages.error(request, "Error updating availability. Please check the form.")
    else:
        form = CounsellorAvailabilityForm(instance=availability)
    return render(request, "My_Admin/CounsellorAvailability/availability_form.html", {"form": form, "title": "Update Availability"})


# Delete record
@csrf_exempt
def availability_delete(request, pk):
    availability = get_object_or_404(CounsellorAvailability, pk=pk)
    if request.method == "POST":
        availability.delete()
        messages.success(request, "Availability deleted successfully!")
        return JsonResponse({"success": True})
    return JsonResponse({"success": False, "error": "Invalid request"})
#####################################################(For Session)################################################################################################
#####################################################(For Session)###################################################################################
class CounsellingSessionForm(forms.ModelForm):
    class Meta:
        model = CounsellingSession
        fields = [
            'appointment', 'session_notes', 'counsellor_observations',
            'recommendations', 'follow_up_required', 'follow_up_date',
            'rating', 'student_feedback'
        ]
        widgets = {
            'session_notes': forms.Textarea(attrs={'class': 'form-control'}),
            'counsellor_observations': forms.Textarea(attrs={'class': 'form-control'}),
            'recommendations': forms.Textarea(attrs={'class': 'form-control'}),
            'student_feedback': forms.Textarea(attrs={'class': 'form-control'}),
            'follow_up_date': forms.DateInput(attrs={'type': 'date', 'class': 'form-control'}),
            'rating': forms.NumberInput(attrs={'class': 'form-control'}),
            'appointment': forms.Select(attrs={'class': 'form-control'}),
        }
# ----------------------------
# CRUD Views
# ----------------------------
def admin_session_list(request):
    sessions = CounsellingSession.objects.all().order_by('-session_date')
    return render(request, "My_Admin/sessions/session_list.html", {"sessions": sessions})


def admin_session_add(request):
    if request.method == "POST":
        form = CounsellingSessionForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Session record added successfully!")
            return redirect("admin_session_list")
        else:
            messages.error(request, "Error adding session. Please check the form.")
    else:
        form = CounsellingSessionForm()
    return render(request, "My_Admin/sessions/session_form.html", {"form": form, "title": "Add Counselling Session"})


def admin_session_update(request, pk):
    session = get_object_or_404(CounsellingSession, pk=pk)
    if request.method == "POST":
        form = CounsellingSessionForm(request.POST, instance=session)
        if form.is_valid():
            form.save()
            messages.success(request, "Session record updated successfully!")
            return redirect("admin_session_list")
        else:
            messages.error(request, "Error updating session.")
    else:
        form = CounsellingSessionForm(instance=session)
    return render(request, "My_Admin/sessions/session_form.html", {"form": form, "title": "Update Counselling Session"})


def admin_session_delete(request, pk):
    session = get_object_or_404(CounsellingSession, pk=pk)
    if request.method == "POST":
        session.delete()
        messages.success(request, "Session deleted successfully!")
        return redirect("admin_session_list")
    return render(request, "My_Admin/sessions/session_confirm_delete.html", {"session": session})
#####################################################(For Session document)################################################################################################
#####################################################(For Session document)###################################################################################
# View all records
def session_document_list(request):
    documents = SessionDocument.objects.all().order_by('-uploaded_at')
    return render(request, 'My_Admin/session_document/session_document_list.html', {'documents': documents})

# Add record
def session_document_add(request):
    if request.method == 'POST':
        form = SessionDocumentForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, "Document added successfully!")
            return redirect('session_document_list')
        else:
            messages.error(request, "Failed to add document. Please correct the errors.")
    else:
        form = SessionDocumentForm()
    return render(request, 'My_Admin/session_document/session_document_form.html', {'form': form, 'title': 'Add Document'})

# Update record
def session_document_edit(request, pk):
    document = get_object_or_404(SessionDocument, pk=pk)
    if request.method == 'POST':
        form = SessionDocumentForm(request.POST, request.FILES, instance=document)
        if form.is_valid():
            form.save()
            messages.success(request, "Document updated successfully!")
            return redirect('session_document_list')
        else:
            messages.error(request, "Failed to update document. Please correct the errors.")
    else:
        form = SessionDocumentForm(instance=document)
    return render(request, 'My_Admin/session_document/session_document_form.html', {'form': form, 'title': 'Edit Document'})

# Delete record
def session_document_delete(request, pk):
    document = get_object_or_404(SessionDocument, pk=pk)
    if request.method == 'POST':
        document.delete()
        messages.success(request, "Document deleted successfully!")
        return redirect('session_document_list')
    return render(request, 'My_Admin/session_document/session_document_confirm_delete.html', {'document': document})
#####################################################(EmergencyConact)################################################################################################
#####################################################(EmergencyConact)###################################################################################
# View all records
def emergency_contact_list(request):
    contacts = EmergencyContact.objects.select_related('student').all()
    return render(request, 'My_Admin/EmergencyContact/emergency_contact_list.html', {'contacts': contacts})

# Add record
def emergency_contact_add(request):
    if request.method == 'POST':
        form = EmergencyContactForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Emergency contact added successfully!")
            return redirect('emergency_contact_list')
        else:
            messages.error(request, "Error adding emergency contact. Please check the form.")
    else:
        form = EmergencyContactForm()
    return render(request, 'My_Admin/EmergencyContact/emergency_contact_form.html', {'form': form, 'title': 'Add Emergency Contact'})

# Update record
def emergency_contact_update(request, pk):
    contact = get_object_or_404(EmergencyContact, pk=pk)
    if request.method == 'POST':
        form = EmergencyContactForm(request.POST, instance=contact)
        if form.is_valid():
            form.save()
            messages.success(request, "Emergency contact updated successfully!")
            return redirect('emergency_contact_list')
        else:
            messages.error(request, "Error updating emergency contact. Please check the form.")
    else:
        form = EmergencyContactForm(instance=contact)
    return render(request, 'My_Admin/EmergencyContact/emergency_contact_form.html', {'form': form, 'title': 'Update Emergency Contact'})

# Delete record
def emergency_contact_delete(request, pk):
    contact = get_object_or_404(EmergencyContact, pk=pk)
    contact.delete()
    messages.success(request, "Emergency contact deleted successfully!")
    return redirect('emergency_contact_list')
#####################################################(CounsellorReview)################################################################################################
#####################################################(CounsellorReview)###################################################################################
# List all reviews
def review_list(request):
    reviews = CounsellorReview.objects.select_related("counsellor__user", "student__user").all()
    return render(request, "My_Admin/reviews/review_list.html", {"reviews": reviews})


# Add new review
def review_add(request):
    counsellors = Counsellor.objects.select_related("user").all()
    students = Student.objects.select_related("user").all()

    if request.method == "POST":
        try:
            counsellor_id = request.POST.get("counsellor")
            student_id = request.POST.get("student")
            rating = request.POST.get("rating")
            review_text = request.POST.get("review_text")

            counsellor = Counsellor.objects.get(pk=counsellor_id)
            student = Student.objects.get(pk=student_id)

            CounsellorReview.objects.create(
                counsellor=counsellor,
                student=student,
                rating=rating,
                review_text=review_text,
            )
            messages.success(request, "Review added successfully!")
            return redirect("review_list")
        except Exception as e:
            messages.error(request, f"Error adding review: {str(e)}")

    return render(request, "My_Admin/reviews/review_add.html", {
        "counsellors": counsellors,
        "students": students,
    })


# Update review
def review_update(request, pk):
    review = get_object_or_404(CounsellorReview, pk=pk)
    counsellors = Counsellor.objects.select_related("user").all()
    students = Student.objects.select_related("user").all()

    if request.method == "POST":
        try:
            review.counsellor_id = request.POST.get("counsellor")
            review.student_id = request.POST.get("student")
            review.rating = request.POST.get("rating")
            review.review_text = request.POST.get("review_text")
            review.save()

            messages.success(request, "Review updated successfully!")
            return redirect("review_list")
        except Exception as e:
            messages.error(request, f"Error updating review: {str(e)}")

    return render(request, "My_Admin/reviews/review_update.html", {
        "review": review,
        "counsellors": counsellors,
        "students": students,
    })
# Delete review
@csrf_exempt
def review_delete(request, pk):
    try:
        review = get_object_or_404(CounsellorReview, pk=pk)
        review.delete()
        return JsonResponse({"success": True, "message": "Review deleted successfully!"})
    except Exception as e:
        return JsonResponse({"success": False, "message": f"Error: {str(e)}"})
#####################################################(report menu)################################################################################################
#####################################################(report menu)###################################################################################
# ---------------- Student Report ----------------
def student_report(request):
    students = Student.objects.all()
    return render(request, "My_Admin/reports/student_report.html", {"students": students})


def export_student_pdf(request):
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="students_report_{timezone.now().date()}.pdf"'

    doc = SimpleDocTemplate(response, pagesize=A4)
    elements = []
    styles = getSampleStyleSheet()

    elements.append(Paragraph("Student Report", styles['Title']))
    elements.append(Spacer(1, 12))

    data = [["ID", "Name", "Faculty", "Department", "Year", "Contact"]]
    for s in Student.objects.all():
        data.append([s.student_id, s.user.get_full_name(), s.get_faculty_display(),
                     s.department, s.academic_year, s.emergency_contact_phone])

    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#007bff")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
    ]))
    elements.append(table)

    doc.build(elements)
    return response


def export_student_excel(request):
    response = HttpResponse(content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = f'attachment; filename="students_report_{timezone.now().date()}.xlsx"'

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Students"

    ws.append(["ID", "Name", "Faculty", "Department", "Year", "Contact"])
    for s in Student.objects.all():
        ws.append([s.student_id, s.user.get_full_name(), s.get_faculty_display(),
                   s.department, s.academic_year, s.emergency_contact_phone])

    wb.save(response)
    return response


# ---------------- Counselling Session Report ----------------
def counselling_report(request):
    sessions = CounsellingSession.objects.all()
    return render(request, "My_Admin/reports/counselling_report.html", {"sessions": sessions})


def export_counselling_pdf(request):
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="counselling_report_{timezone.now().date()}.pdf"'

    doc = SimpleDocTemplate(response, pagesize=A4)
    elements = []
    styles = getSampleStyleSheet()
    elements.append(Paragraph("Counselling Sessions Report", styles['Title']))
    elements.append(Spacer(1, 12))

    data = [["Student", "Counsellor", "Date", "Follow-up", "Rating"]]
    for session in CounsellingSession.objects.select_related("appointment__student", "appointment__counsellor"):
        data.append([
            session.appointment.student.user.get_full_name(),
            session.appointment.counsellor.user.get_full_name(),
            session.session_date.strftime("%Y-%m-%d %H:%M"),
            "Yes" if session.follow_up_required else "No",
            session.rating or "N/A"
        ])

    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.green),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
    ]))
    elements.append(table)

    doc.build(elements)
    return response


def export_counselling_excel(request):
    response = HttpResponse(content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = f'attachment; filename="counselling_report_{timezone.now().date()}.xlsx"'

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Counselling Sessions"

    ws.append(["Student", "Counsellor", "Date", "Follow-up", "Rating"])
    for session in CounsellingSession.objects.select_related("appointment__student", "appointment__counsellor"):
        ws.append([
            session.appointment.student.user.get_full_name(),
            session.appointment.counsellor.user.get_full_name(),
            session.session_date.strftime("%Y-%m-%d %H:%M"),
            "Yes" if session.follow_up_required else "No",
            session.rating or "N/A"
        ])

    wb.save(response)
    return response


# ---------------- Appointment Report ----------------
def appointment_report(request):
    appointments = Appointment.objects.all()
    return render(request, "My_Admin/reports/appointment_report.html", {"appointments": appointments})


def export_appointment_pdf(request):
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="appointment_report_{timezone.now().date()}.pdf"'

    doc = SimpleDocTemplate(response, pagesize=A4)
    elements = []
    styles = getSampleStyleSheet()
    elements.append(Paragraph("Appointment Report", styles['Title']))
    elements.append(Spacer(1, 12))

    data = [["Student", "Counsellor", "Date", "Status", "Purpose"]]
    for a in Appointment.objects.select_related("student", "counsellor"):
        data.append([
            a.student.user.get_full_name(),
            a.counsellor.user.get_full_name(),
            a.appointment_date.strftime("%Y-%m-%d %H:%M"),
            a.get_status_display(),
            a.purpose[:30] + "..." if len(a.purpose) > 30 else a.purpose
        ])

    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
    ]))
    elements.append(table)

    doc.build(elements)
    return response


def export_appointment_excel(request):
    response = HttpResponse(content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = f'attachment; filename="appointment_report_{timezone.now().date()}.xlsx"'

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Appointments"

    ws.append(["Student", "Counsellor", "Date", "Status", "Purpose"])
    for a in Appointment.objects.select_related("student", "counsellor"):
        ws.append([
            a.student.user.get_full_name(),
            a.counsellor.user.get_full_name(),
            a.appointment_date.strftime("%Y-%m-%d %H:%M"),
            a.get_status_display(),
            a.purpose
        ])

    wb.save(response)
    return response
#####################################################(roles permission menu)################################################################################################
#####################################################(roles permission menu)###################################################################################
# -----------------------------
# ROLE & PERMISSION MANAGEMENT
# -----------------------------

@login_required
@permission_required('auth.view_group', raise_exception=True)
def role_list(request):
    roles = Group.objects.all()
    return render(request, "My_Admin/roles/role_list.html", {"roles": roles})

@login_required
@permission_required('auth.add_group', raise_exception=True)
def role_add(request):
    if request.method == "POST":
        role_name = request.POST.get("name")
        if Group.objects.filter(name=role_name).exists():
            messages.error(request, "Role already exists!")
        else:
            Group.objects.create(name=role_name)
            messages.success(request, "Role created successfully!")
        return redirect("role_list")
    return render(request, "My_Admin/roles/role_add.html")

@login_required
@permission_required('auth.change_group', raise_exception=True)
def role_edit(request, pk):
    role = get_object_or_404(Group, pk=pk)
    if request.method == "POST":
        role.name = request.POST.get("name")
        role.save()
        messages.success(request, "Role updated successfully!")
        return redirect("role_list")
    return render(request, "My_Admin/roles/role_edit.html", {"role": role})

@login_required
@permission_required('auth.delete_group', raise_exception=True)
def role_delete(request, pk):
    role = get_object_or_404(Group, pk=pk)
    role.delete()
    messages.success(request, "Role deleted successfully!")
    return redirect("role_list")

# -----------------------------
# SYSTEM SETTINGS (Example: Manage Users)
# -----------------------------

@login_required
@permission_required('My_Admin.can_manage_users', raise_exception=True)
def user_list(request):
    users = CustomUser.objects.all()
    return render(request, "My_Admin/settings/user_list.html", {"users": users})

@login_required
@permission_required('My_Admin.can_manage_users', raise_exception=True)
def user_add(request):
    if request.method == "POST":
        username = request.POST.get("username")
        email = request.POST.get("email")
        role = request.POST.get("role")
        user = CustomUser.objects.create_user(
            username=username,
            email=email,
            password="default123",
            role=role
        )
        messages.success(request, "User added successfully!")
        return redirect("user_list")
    return render(request, "My_Admin/settings/user_add.html")

@login_required
@permission_required('My_Admin.can_manage_users', raise_exception=True)
def user_edit(request, pk):
    user = get_object_or_404(CustomUser, pk=pk)
    if request.method == "POST":
        user.username = request.POST.get("username")
        user.email = request.POST.get("email")
        user.role = request.POST.get("role")
        user.save()
        messages.success(request, "User updated successfully!")
        return redirect("user_list")
    return render(request, "My_Admin/settings/user_edit.html", {"user": user})

@login_required
@permission_required('My_Admin.can_manage_users', raise_exception=True)
def user_delete(request, pk):
    user = get_object_or_404(CustomUser, pk=pk)
    user.delete()
    messages.success(request, "User deleted successfully!")
    return redirect("user_list")
#####################################################################################################################################################
#####################################################################################################################################################
@login_required
def student_dashboard(request):
    if not request.user.is_student():
        messages.error(request, 'Access denied. Student account required.')
        return redirect('login')
    
    try:
        # Get current time with timezone awareness
        now = timezone.now()
        
        # Check if user has a student profile, create one if it doesn't exist
        if not hasattr(request.user, 'student'):
            # Create student profile if it doesn't exist
            from .models import Student
            student = Student.objects.create(
                user=request.user,
                student_id=f"STU{request.user.id:06d}",
                enrollment_date=timezone.now().date(),
                faculty='engineering',  # Default faculty
                department='Computer Science',  # Default department
                academic_year=1,  # Default year
                emergency_contact_name='Parent/Guardian',
                emergency_contact_phone='000-000-0000'
            )
            messages.info(request, 'Student profile has been created for you.')
        else:
            student = request.user.student
        
        # Get statistics using the models directly from current app
        # Count upcoming appointments
        try:
            from .models import Appointment
            upcoming_appointments = Appointment.objects.filter(
                student=student,
                appointment_date__gte=now,
                status__in=['pending', 'confirmed']
            ).count()
            
            # Get upcoming appointments list
            upcoming_appointments_list = Appointment.objects.filter(
                student=student,
                appointment_date__gte=now,
                status__in=['pending', 'confirmed']
            ).select_related('counsellor', 'counsellor__user').order_by('appointment_date')[:5]
            
            total_appointments = Appointment.objects.filter(student=student).count()
            completed_sessions = Appointment.objects.filter(student=student, status='completed').count()
            
        except Exception as e:
            # If Appointment model doesn't exist or there's an error, use dummy data
            upcoming_appointments = 0
            upcoming_appointments_list = []
            total_appointments = 0
            completed_sessions = 0
        
        # Count available counsellors
        try:
            from .models import Counsellor
            available_counsellors = Counsellor.objects.filter(is_available=True).count()
        except:
            available_counsellors = 3  # Default value
        
        # Get recent activity (using appointments as sessions)
        try:
            from .models import Appointment
            recent_sessions = Appointment.objects.filter(
                student=student
            ).select_related('counsellor', 'counsellor__user').order_by('-appointment_date')[:5]
        except:
            recent_sessions = []
        
        # Set default values for features
        pending_reviews = 0
        submitted_reviews = 0
        documents_count = 0
        total_sessions = completed_sessions
        
        context = {
            'upcoming_appointments': upcoming_appointments,
            'completed_sessions': completed_sessions,
            'pending_reviews': pending_reviews,
            'available_counsellors': available_counsellors,
            'upcoming_appointments_list': upcoming_appointments_list,
            'recent_sessions': recent_sessions,
            'total_appointments': total_appointments,
            'total_sessions': total_sessions,
            'submitted_reviews': submitted_reviews,
            'documents_count': documents_count,
            'student': student,
        }
        
        return render(request, 'My_Student/student_dashboard.html', context)
        
    except Exception as e:
        messages.error(request, f'Error loading dashboard: {str(e)}')
        # Return a basic dashboard even if there are errors
        return render(request, 'My_Student/student_dashboard.html', {
            'upcoming_appointments': 0,
            'completed_sessions': 0,
            'pending_reviews': 0,
            'available_counsellors': 0,
            'upcoming_appointments_list': [],
            'recent_sessions': [],
            'total_appointments': 0,
            'total_sessions': 0,
            'submitted_reviews': 0,
            'documents_count': 0,
            'student': None,
        })

#####################################################################################################################################################
# View all appointments
def student_appointment_list(request):
    appointments = Appointment.objects.all().select_related('student', 'counsellor')
    return render(request, 'My_Student/appointments/list.html', {'appointments': appointments})

# Add new appointment
def Student_appointment_create(request):
    if request.method == "POST":
        form = AppointmentForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Appointment created successfully!")
            return redirect('student_appointment_list')
        else:
            messages.error(request, "Error creating appointment. Please check the form.")
    else:
        form = AppointmentForm()
    return render(request, 'My_Student/appointments/form.html', {'form': form, 'title': 'Add Appointment'})
#####################################################################################################################################################
# ----------------------------
def student_session_list(request):
    sessions = CounsellingSession.objects.all().order_by('-session_date')
    return render(request, "My_Student/sessions/session_list.html", {"sessions": sessions})
#################################################################################################################
# View all records
def student_session_document_list(request):
    documents = SessionDocument.objects.all().order_by('-uploaded_at')
    return render(request, 'My_Student/session_document/session_document_list.html', {'documents': documents})
##################################################################################################################
# View all records
def student_emergency_contact_list(request):
    contacts = EmergencyContact.objects.select_related('student').all()
    return render(request, 'My_Student/EmergencyContact/emergency_contact_list.html', {'contacts': contacts})
####################################################################################################################
# List all reviews
def student_review_list(request):
    reviews = CounsellorReview.objects.select_related("counsellor__user", "student__user").all()
    return render(request, "My_Student/reviews/review_list.html", {"reviews": reviews})


# Add new review
def student_review_add(request):
    counsellors = Counsellor.objects.select_related("user").all()
    students = Student.objects.select_related("user").all()

    if request.method == "POST":
        try:
            counsellor_id = request.POST.get("counsellor")
            student_id = request.POST.get("student")
            rating = request.POST.get("rating")
            review_text = request.POST.get("review_text")

            counsellor = Counsellor.objects.get(pk=counsellor_id)
            student = Student.objects.get(pk=student_id)

            CounsellorReview.objects.create(
                counsellor=counsellor,
                student=student,
                rating=rating,
                review_text=review_text,
            )
            messages.success(request, "Review added successfully!")
            return redirect("student_review_list")
        except Exception as e:
            messages.error(request, f"Error adding review: {str(e)}")

    return render(request, "My_Student/reviews/review_add.html", {
        "counsellors": counsellors,
        "students": students,
    })


# Update review
def student_review_update(request, pk):
    review = get_object_or_404(CounsellorReview, pk=pk)
    counsellors = Counsellor.objects.select_related("user").all()
    students = Student.objects.select_related("user").all()

    if request.method == "POST":
        try:
            review.counsellor_id = request.POST.get("counsellor")
            review.student_id = request.POST.get("student")
            review.rating = request.POST.get("rating")
            review.review_text = request.POST.get("review_text")
            review.save()

            messages.success(request, "Review updated successfully!")
            return redirect("student_review_list")
        except Exception as e:
            messages.error(request, f"Error updating review: {str(e)}")

    return render(request, "My_Student/reviews/review_update.html", {
        "review": review,
        "counsellors": counsellors,
        "students": students,
    })
#############################################################################################################
@login_required
def student_user_list(request):
    # Allow only Student role
    if not request.user.is_student():
        raise PermissionDenied("You are not allowed to access this page.")

    users = CustomUser.objects.filter(role="student")  # fetch only student users
    return render(request, "My_Student/settings/user_list.html", {"users": users})
#####################################################################################################################################################
@login_required
def counsellor_dashboard(request):
    if not hasattr(request.user, 'counsellor'):
        messages.error(request, 'Access denied. Counsellor account required.')
        return redirect('login')
    
    # Get counsellor profile
    counsellor = request.user.counsellor
    
    # Dashboard statistics
    today = timezone.now().date()
    
    # Today's appointments
    today_appointments = Appointment.objects.filter(
        counsellor=counsellor,
        appointment_date__date=today
    ).count()
    
    # Pending appointments
    pending_appointments = Appointment.objects.filter(
        counsellor=counsellor,
        status='pending'
    ).count()
    
    # Recent reviews (last 30 days)
    recent_reviews = CounsellorReview.objects.filter(
        counsellor=counsellor,
        created_at__gte=timezone.now() - timedelta(days=30)
    ).count()
    
    # Active students (students with appointments in last 30 days)
    active_students = Student.objects.filter(
        appointment__counsellor=counsellor,
        appointment__appointment_date__gte=timezone.now() - timedelta(days=30)
    ).distinct().count()
    
    # Additional statistics
    total_appointments = Appointment.objects.filter(counsellor=counsellor).count()
    
    # Completed sessions - using appointment status instead of session status
    completed_sessions = CounsellingSession.objects.filter(
        appointment__counsellor=counsellor,
        appointment__status='completed'
    ).count()
    
    # Average rating
    reviews = CounsellorReview.objects.filter(counsellor=counsellor)
    average_rating = reviews.aggregate(Avg('rating'))['rating__avg'] or 0.0
    
    # Documents count
    documents_count = SessionDocument.objects.filter(
        session__appointment__counsellor=counsellor
    ).count()
    
    # Today's appointments list
    today_appointments_list = Appointment.objects.filter(
        counsellor=counsellor,
        appointment_date__date=today
    ).select_related('student__user').order_by('appointment_date')[:5]
    
    context = {
        'counsellor': counsellor,
        'today_appointments': today_appointments,
        'pending_appointments': pending_appointments,
        'recent_reviews': recent_reviews,
        'active_students': active_students,
        'total_appointments': total_appointments,
        'completed_sessions': completed_sessions,
        'average_rating': round(average_rating, 1),
        'documents_count': documents_count,
        'today_appointments_list': today_appointments_list,
    }
    
    return render(request, 'My_Consuller/counsellor_dashboard.html', context)
#####################################################################################################################################################
#####################################################################################################################################################
# View all appointments
def counseller_appointment_list(request):
    appointments = Appointment.objects.all().select_related('student', 'counsellor')
    return render(request, 'My_Consuller/appointments/list.html', {'appointments': appointments})
################################################################################################################
# List all availabilities (for logged-in counsellor)
@login_required
@permission_required('My_Appointment.can_manage_own_availability', raise_exception=True)
def manage_availability(request):
    availabilities = CounsellorAvailability.objects.filter(counsellor__user=request.user)
    return render(request, "My_Consuller/appointments/manage_availability.html", {
        "availabilities": availabilities
    })
################################################################################################################
# Add availability
@login_required
@permission_required('My_Appointment.can_manage_own_availability', raise_exception=True)
def add_availability(request):
    if request.method == "POST":
        form = CounsellorAvailabilityForm(request.POST)
        if form.is_valid():
            availability = form.save(commit=False)
            availability.counsellor = request.user.counsellor  # assuming request.user has counsellor profile
            availability.save()
            messages.success(request, "Availability added successfully!")
            return redirect("manage_availability")
        else:
            messages.error(request, "Error adding availability. Please check the form.")
    else:
        form = CounsellorAvailabilityForm()
    return render(request, "My_Consuller/appointments/add_availability.html", {"form": form})
################################################################################################################
# Update availability
@login_required
@permission_required('My_Appointment.can_manage_own_availability', raise_exception=True)
def update_availability(request, pk):
    availability = get_object_or_404(CounsellorAvailability, pk=pk, counsellor__user=request.user)
    if request.method == "POST":
        form = CounsellorAvailabilityForm(request.POST, instance=availability)
        if form.is_valid():
            form.save()
            messages.success(request, "Availability updated successfully!")
            return redirect("manage_availability")
        else:
            messages.error(request, "Error updating availability.")
    else:
        form = CounsellorAvailabilityForm(instance=availability)
    return render(request, "My_Consuller/appointments/update_availability.html", {"form": form})
################################################################################################################
# Delete availability
@login_required
@permission_required('My_Appointment.can_manage_own_availability', raise_exception=True)
def delete_availability(request, pk):
    availability = get_object_or_404(CounsellorAvailability, pk=pk, counsellor__user=request.user)
    availability.delete()
    messages.success(request, "Availability deleted successfully!")
    return redirect("manage_availability")
################################################################################################################
# View availability (readonly for others)
@login_required
def view_availability(request):
    availabilities = CounsellorAvailability.objects.filter(is_active=True)
    return render(request, "My_Consuller/appointments/view_availability.html", {
        "availabilities": availabilities
    })
################################################################################################################################################################
@login_required
@permission_required('My_Session.can_view_session_history', raise_exception=True)
def session_list(request):
    sessions = CounsellingSession.objects.all().order_by('-session_date')
    return render(request, 'My_Consuller/My_Session/session_list.html', {'sessions': sessions})


@login_required
@permission_required('My_Session.can_write_session_notes', raise_exception=True)
def session_create(request):
    if request.method == "POST":
        form = CounsellingSessionForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Session added successfully!")
            return redirect('my_sessions_list')
        else:
            messages.error(request, "Error adding session. Please check the form.")
    else:
        form = CounsellingSessionForm()
    return render(request, 'My_Consuller/My_Session/session_form.html', {'form': form, 'title': 'Add Session'})


@login_required
@permission_required('My_Session.can_write_session_notes', raise_exception=True)
def session_update(request, pk):
    session = get_object_or_404(CounsellingSession, pk=pk)
    if request.method == "POST":
        form = CounsellingSessionForm(request.POST, instance=session)
        if form.is_valid():
            form.save()
            messages.success(request, "Session updated successfully!")
            return redirect('my_sessions_list')
        else:
            messages.error(request, "Error updating session. Please check the form.")
    else:
        form = CounsellingSessionForm(instance=session)
    return render(request, 'My_Consuller/My_Session/session_form.html', {'form': form, 'title': 'Update Session'})


@login_required
@permission_required('My_Session.can_write_session_notes', raise_exception=True)
def session_delete(request, pk):
    session = get_object_or_404(CounsellingSession, pk=pk)
    session.delete()
    messages.success(request, "Session deleted successfully!")
    return redirect('my_sessions_list')
###########################################################################################################################################################
# View all records
def counseller_session_document_list(request):
    documents = SessionDocument.objects.all().order_by('-uploaded_at')
    return render(request, 'My_Consuller/session_document/session_document_list.html', {'documents': documents})

# Add record
def counseller_session_document_add(request):
    if request.method == 'POST':
        form = SessionDocumentForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, "Document added successfully!")
            return redirect('counseller_session_document_list')
        else:
            messages.error(request, "Failed to add document. Please correct the errors.")
    else:
        form = SessionDocumentForm()
    return render(request, 'My_Consuller/session_document/session_document_form.html', {'form': form, 'title': 'Add Document'})

# Update record
def counseller_session_document_edit(request, pk):
    document = get_object_or_404(SessionDocument, pk=pk)
    if request.method == 'POST':
        form = SessionDocumentForm(request.POST, request.FILES, instance=document)
        if form.is_valid():
            form.save()
            messages.success(request, "Document updated successfully!")
            return redirect('counseller_session_document_list')
        else:
            messages.error(request, "Failed to update document. Please correct the errors.")
    else:
        form = SessionDocumentForm(instance=document)
    return render(request, 'My_Consuller/session_document/session_document_form.html', {'form': form, 'title': 'Edit Document'})
########################################################################################################################################
@login_required
def counseller_student_list(request):
    students = Student.objects.all()
    return render(request, 'My_Consuller/Student_Management/student_list.html', {'students': students})
#################################################################################################################
# List all reviews
def counseller_review_list(request):
    reviews = CounsellorReview.objects.select_related("counsellor__user", "student__user").all()
    return render(request, "My_Consuller/reviews/review_list.html", {"reviews": reviews})
#######################################################################################################################
@login_required
def counseller_user_list(request):
    # Allow only Student role
    if not request.user.is_counsellor():
        raise PermissionDenied("You are not allowed to access this page.")

    users = CustomUser.objects.filter(role="counsellor")  # fetch only student users
    return render(request, "My_Consuller/settings/user_list.html", {"users": users})